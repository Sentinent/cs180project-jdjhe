import xlrd
import pandas as pd
import openpyxl
import xlwt
from xlwt import Workbook
import csv

#from data_structures.violation_objects import *
from violation_objects import *



class Parsing:
    # Function to parse the total amount of Violations occuring in the data sheet
    def ParseViolations(self):
        violationsList = []

        """
        Set the sise of the list to 99 by appending 99 0's to it
        This is done so that the index of the list represents the Violation Code from the data set
        """
        for zero in range(99):
            violationsList.append(0)

        with open('../datasets/parking-violations-issued-fiscal-year-2014-august-2013-june-2014.csv', encoding='utf-8', newline='') as file:
            reader = enumerate(csv.reader(file))

            """
            Read through the data and using the data with a 
            given data cell as the index for the list to increment 
            its violation count
            """
            for i, row in reader:
                if i > 0:
                    var = row[5]
                    violationsList[int(var) - 1] += 1

        # Creating a workbook object in order to save the new data
        wb = Workbook()
        sheet1 = wb.add_sheet('Sheet 1')

        sheet1.write(0, 0, 'Violation Code')
        sheet1.write(0, 1, 'Counts')

        # Write all of the saved data to the workbook
        index = 1
        for v in violationsList:
            sheet1.write(index, 0, index)
            sheet1.write(index, 1, v)
            index += 1

        # Save the new workbook as a file
        wb.save('../parsed_data/violationCounts.xls')
    
    # Function to collect the amount of times an object has had a violation
    def ObjectViolations(self, col, objectName, saveFile):
        list = []
        with open('../datasets/parking-violations-issued-fiscal-year-2014-august-2013-june-2014.csv', encoding='utf-8', newline='') as file:
            reader = enumerate(csv.reader(file))

            """
            If the data in the given cell from the data sheet is in the list
            then simply add 1 to its violation count
            otherwise add it to the list
            """
            check = 1
            for i, row in reader:
                if (i > 0):
                    for val in list:
                        if (val.object == row[col]):
                            val.violation += 1
                            check = 0
                            break
                    if (check):
                        list.append(ViolationObject(1, row[col]))
                    check = 1

            # Set up a workbook in order to save the data
            wb = Workbook()
            sheet1 = wb.add_sheet('Sheet 1')

            sheet1.write(0, 0, objectName)
            sheet1.write(0, 1, 'Violation Counts')

            # Write the data to a file to be saved
            index = 1
            for v in list:
                sheet1.write(index, 0, v.object)
                sheet1.write(index, 1, v.violation)
                index += 1

            # Save the new file
            wb.save(f"../parsed_data/{saveFile}.xls")


    # Function to get the types of violations with their respective counts with another object associated with it such as counties
    def ListObjectViolations(self, col, objectName, saveFile):
        list = []
        violationsList = []
        with open('../datasets/parking-violations-issued-fiscal-year-2014-august-2013-june-2014.csv', encoding='utf-8', newline='') as file:
            reader = enumerate(csv.reader(file))

            """
            Read through the desired Data Column and if we encounter a new value add it to the list
            If the value is already in the list, then find the associated violation for that value 
            and add it to the value's sublist of violations with their counts
            """
            check = 1
            for i, row in reader:
                if (i > 0):
                    for val in list:
                        if (val.primaryObject == row[col]):
                            val.addListing(row[5])
                            check = 0
                            break
                    if (check == 1):
                        list.append(ObjectListViolations(row[col]))
                    check = 1 

            # Create a workbook in order to save the data to another file
            wb = Workbook()
            sheet1 = wb.add_sheet('Sheet 1')

            sheet1.write(0, 0, objectName)

            print(list[0].getListing(0))

            prime_index = 1
            for v in list:
                sheet1.write(prime_index, 0, v.primaryObject)
                sheet1.write(prime_index + 1, 0, "Violation Code")
                sheet1.write(prime_index + 2, 0, "Violation Count")
                index = 1
                for j in v.list:
                    sheet1.write(prime_index + 1, index, int(j.object))
                    sheet1.write(prime_index + 2, index, j.violation)
                    index += 1
                prime_index += 3

            wb.save(f"../parsed_data/{saveFile}.xls")

    # Function to parse data to collect the number of violations occuring at a given time
    def TimeViolations(self, objectName, saveFile):
        list = []
        with open('../datasets/parking-violations-issued-fiscal-year-2014-august-2013-june-2014.csv', encoding='utf-8', newline='') as file:
            reader = enumerate(csv.reader(file))

            # Read through the time column
            # Save the value as a string then check to see if it is in the form 0000A
            # If the first character is either 0 or 1 then it is a valid time entry, so save the hour as a string variable
            # Otherwise set the string variable to "Unclear Time"
            # Then store it in the list if it is not present, while if it is there then add 1 to the violation count 
            time = ""
            check = 1
            for i, row in reader:
                if (i > 0):
                    timeWord = str(row[19])
                    if (len(timeWord) >= 5):
                        if (timeWord[0] == "1" or timeWord[0] == "0"):
                            time = timeWord[0] + timeWord[1] + timeWord[4] + "M"
                        else:
                            time = "Unclear Time"
                    else:
                        time = "Unclear Time"

                    true = findObject(list, time)
                    if (true != -1):
                        list[true].violation += 1
                    else:
                        list.append(ViolationObject(1, time))

            # Open a workbook to save the data
            wb = Workbook()
            sheet1 = wb.add_sheet('Sheet 1')

            sheet1.write(0, 0, objectName)
            sheet1.write(0, 1, 'Violation Counts')

            # Read through the list of times and write them to a file
            index = 1
            for v in list:
                sheet1.write(index, 0, v.object)
                sheet1.write(index, 1, v.violation)
                index += 1

            # Save the file
            wb.save(f"../parsed_data/{saveFile}.xls") 
